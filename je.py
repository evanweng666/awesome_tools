import sys
import os
import time
import random
import openpyxl

"""
@Title  excel和json互转工具
@Author Evan
"""


def json_to_excel_new(data, filename=None, sheet_name=None, rules=None):
    """
    json转excel - 新建/覆盖
    :param data: json数据（[[1, 2, 3], [4, 5, 6]]）
    :param filename: excel文件路径（格式要求：xlsx）
    :param sheet_name: 指定工作表（默认第一个）
    :param rules: [[0, 0, 0, 1], [1, 0, 1, 1]]，单元格合并规则[start_row, start_column, end_row, end_column]
    :return:
    """
    if not filename:
        filename = '{}{}{}_{}.xlsx'.format(sys.path[0], '\\', time.strftime("%Y%m%d", time.localtime()),
                                           ''.join(random.sample('0123456789', 4)))
    elif not filename.endswith('.xlsx'):
        raise Exception('The type of file should be xlsx.')
    book = openpyxl.Workbook()
    sheet = book.active
    if sheet_name:
        sheet.title = sheet_name
    if data:
        for i in range(len(data)):
            row = data[i]
            for j in range(len(row)):
                sheet.cell(i + 1, j + 1, str(row[j]))
    if rules:
        for rule in rules:
            if len(rule) != 4:
                raise Exception('The rule {} is wrong.'.format(rule))
            else:
                sheet.merge_cells(start_row=rule[0] + 1, start_column=rule[1] + 1, end_row=rule[2] + 1,
                                  end_column=rule[3] + 1)
    book.save(filename)


def json_to_excel_append(data, filename, sheet_name=None, rules=None):
    """
    json转excel - 新建/追加
    :param data: json数据（[[1, 2, 3], [4, 5, 6]]）
    :param filename: excel文件路径（格式要求：xlsx）
    :param sheet_name: 指定工作表（默认第一个）
    :param rules: [[0, 0, 0, 1], [1, 0, 1, 1]]，单元格合并规则[start_row, start_column, end_row, end_column]
    :return:
    """
    if not filename:
        raise Exception('The filename can not be empty.')
    if not filename.endswith('.xlsx'):
        raise Exception('The type of file should be xlsx.')
    if not os.path.isfile(filename):
        json_to_excel_new(data, filename, sheet_name, rules)
    else:
        book = openpyxl.load_workbook(filename)
        if sheet_name:
            if book.__contains__(sheet_name):
                sheet = book[sheet_name]
            else:
                sheet = book.create_sheet(sheet_name)
        else:
            sheet = book.active
        # 统计现有行数
        rs = (r for r in sheet.rows)
        cr = sum(1 for _ in rs)
        if data:
            for i in range(len(data)):
                row = data[i]
                for j in range(len(row)):
                    sheet.cell(cr + i + 1, j + 1, str(row[j]))
        if rules:
            for rule in rules:
                if len(rule) != 4:
                    raise Exception('The rule {} is wrong.'.format(rule))
                else:
                    sheet.merge_cells(start_row=cr + rule[0] + 1, start_column=rule[1] + 1, end_row=cr + rule[2] + 1,
                                      end_column=rule[3] + 1)
        book.save(filename)


def excel_to_json(filename, sheet_name=None):
    """
    excel转json
    :param filename: excel文件路径（格式要求：xlsx）
    :param sheet_name: 指定工作表（默认第一个）
    :return: [[1, 2, 3], [4, 5, 6]]
    """
    if not filename:
        raise Exception('The filename can not be empty.')
    if not filename.endswith('.xlsx'):
        raise Exception('The type of file should be xlsx.')
    book = openpyxl.load_workbook(filename)
    if sheet_name:
        if not book.__contains__(sheet_name):
            raise Exception('The sheet {} is not exist.'.format(sheet_name))
        else:
            sheet = book[sheet_name]
    else:
        sheet = book.active
    data = []
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            if cell.value:
                row_data.append(cell.value)
            else:
                row_data.append('')
        data.append(row_data)
    return data


if __name__ == '__main__':
    json_data = [[1, '', 3], [4, '', 6]]
    merge_rules = [[0, 0, 0, 1], [1, 0, 1, 1]]
    json_to_excel_append(json_data, 'test.xlsx', 'Sheet1', merge_rules)
    print(excel_to_json('test.xlsx', 'Sheet1'))
